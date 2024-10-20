import requests
import json
import base64
import pandas as pd
from openpyxl import Workbook, load_workbook
import random
import re

BUFFER_SIZE = 1
VAGUE_SIZE = 4
FILTER_SIZE = 1
url = "https://flag.smarttrot.com/v1/chat/completions"
api_secret_key = ''

headers = {
    "Content-Type": "application/json",
    'Accept': 'application/json',
    'Authorization': "Bearer " + api_secret_key
}

# "role": "assistant"/"user"
assistant = "assistant"
user = "user"

data = {
    "max_tokens": 1200,
    "model": "gpt-4o-2024-05-13",
    "temperature": 0.8,
    "top_p": 1,
    "presence_penalty": 1,
    "messages": [
        # {
        #     "role": "assistant",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "I'm going to give you 1 examples to learn from.Example1: Description:The image shows a scene with a white table on which a blue cube and an open cardboard box are placed. The blue cube is situated on the lower left side of the image, resting on the white surface of the table. It appears to be a solid object with a noticeable shadow indicating a light source from the upper right side.The cardboard box is positioned on the table, closer to the top right corner of the image. The box is open, with its flaps standing upright, and it seems to be empty. The cardboard has a typical brown color with visible creases at the edges where the flaps have been folded.In the background, there's a radiator against the wall and some cables that can be seen hanging loosely, suggesting this might be a workspace or an area where some sort of work or activity is being carried out.The overall scenario seems to be set up for an activity, possibly for the purpose of demonstrating or testing the functionality of a robot arm in picking and placing objects, considering your mention of a two-finger claw robot arm. The blue cube could be the object that needs to be manipulated, and the cardboard box could be the target location for placing the object. Object list: [blue_cube, cardboard_box].  Generated task instruction:Put the blue_cube in the cardboard_box. Is it a vague instruction:No.  Atomic action after disassembly:[(grab,blue_cube),(place,cardboard_box)]."
        #         },
        #         {
        #             "type": "image_url",
        #             "image_url": {
        #                 "url": f"data:image/jpeg;base64,{base64_image_box}"
        #             }
        #         }
        #     ]
        # },
        # {
        #     "role": "user",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "Please complete the task according to the example."
        #         }
        #     ]
        # }
    ]
}


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def read_excel_data(file_path, min_row, max_row):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    names = []
    descriptions = []
    place_methods = []
    directions = []
    # result = ""
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        # print(row[n])
        names.append(row[0])
        descriptions.append(row[1])
        place_methods.append(row[2])
        directions.append(row[3])
        # result = row[n]

    wb.close()
    return names, descriptions, place_methods, directions


def get_example(file_path):
    min_row = 2
    max_row = 47
    names, descriptions, place_methods, directions = read_excel_data(file_path, min_row, max_row)
    res = ""
    for i in range(min_row, max_row + 1):
        name = "Example" + str(i - 1) + ": Name:" + names[i-2] + ","
        description = "Description:" + descriptions[i-2] + ","
        place_method = "放置方式:" + place_methods[i-2] + ","
        direction = "Direction of grasp:" + directions[i-2] + "."
        res = res + name + description + place_method + direction
    return res


def get_text_image(role, txt, base64_image):
    base64_image = encode_image(base64_image)
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            },
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{base64_image}"
                }
            }
        ]
    }

    return res


def get_text(role, txt):
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            }
        ]
    }

    return res


def consult_GPT(question):
    response = requests.post(url, headers=headers, data=json.dumps(question))
    result = response
    return result


def output_format(dict_data):
    for key, value in dict_data.items():
        if key == 'messages':
            print(f"{key}:")
            for item in dict_data['messages']:
                role = item['role']
                print(f"Role: {role}")
                for content_item in item['content']:
                    if isinstance(content_item, dict):
                        if content_item['type'] == 'text':
                            print(f"  Text: {content_item['text']}")
                        elif content_item['type'] == 'image_url':
                            print("  Image url: [NOT DISPLAYED]")
                        else:
                            print(f"  {content_item}")
                    else:
                        print(f"  {item['content']}")
                        break
        else:
            print(f"{key}: {value}")

def get_name(file_path,i):
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active
    row = sheet[i]
    image, command1 = row[0].value, row[2].value
    wb.close()
    return image, command1


def write_to_cell(file_path, row_num, col_num, value):
    wb = load_workbook(file_path)
    sheet = wb.active
    sheet.cell(row=row_num, column=col_num, value=value)
    wb.save(file_path)
    wb.close()


def extract_first_obj(input_string):
    # 去掉开头和结尾的方括号
    input_string = input_string.strip('[]')

    # 分割字符串，提取第一个括号内的内容
    parts = input_string.split('),(')
    if parts:
        first_pair = parts[0].strip('()')
        # 分割第一个括号内的内容，提取物体名称
        obj = first_pair.split(',')[1].strip()
        return obj
    return None

def convert_to_tuple(input_str):
    # 去掉字符串的起始和结束方括号以及两端的空白字符
    trimmed_str = input_str.strip()[1:-1]
    # 按逗号分隔字符串
    split_str = trimmed_str.split(',')
    # 将分割后的列表转换成元组
    result_tuple = tuple(split_str)
    return result_tuple


# command2_dismantle_txt="三级指令是对具有明确物体名称的物体进行明确操作的指令，同时三级指令必须具有一个明确的放置位置（放置别的具体物体的左右前后上边），如将苹果放置到书本上。针对该图片的一级指令是{}，请你将其分解为三级指令。步骤：1、结合图片分析当前场景，并输出当前场景的详细描述（包括有什么物体、物体的位置、物体之间的空间关系）。2、根据图片和场景描述，结合世界通用知识对图中的物体的特征进行详细描述（包括长、宽、高、形状、颜色、位置、与其他物体的空间关系等特征），在该步骤中输出物体及其描述。3、根据一级指令、场景描述、物体描述，分析并输出最终需要实现的场景的描述（包括物体如何摆放）。4、结合图片从一级指令中推理出现在需要对哪些物体进行操作,在该步骤中输出需要操作的物体。5、三级指令是对具有明确物体名称的物体进行明确操作的指令，同时三级指令必须具有一个明确的放置位置（放置别的具体物体的左右前后上边），如将苹果放置到书本上。根据一级指令及其意图，为每个需要被操作的物体生成一个明确的三级指令，输出三级指令。6、结合图片和通用世界知识，确定每个三级指令执行的顺序，使得整个过程流畅且合理。输出经过排序后的三级指令。输出格式为[command1,command2,command3.....],command1代指分解出的三级指令，严格按照格式输出。请你一步一步思考。"
# extract_command3_que="请从以上描述中提取出三级指令,不需要对描述中的三级指令做任何更改，输出格式为：“[command1,command2,command3....]”,严格按照输出格式输出([]内不需要使用引号，注意一定要使用,而不是，)，command1只是代号，并不需要输出，不需要输出其他内容"
command2_dismantle_txt="A level-three command is an instruction to perform a specific action on an object with a clear name, and it must include a precise placement location (e.g., left, right, front, back, or above another specific object), such as placing an apple on a book. The level-one command for this image is {}, please break it down into level-three commands. Steps: 1. Analyze the current scene based on the image and output a detailed description of the scene (including the objects present, their positions, and the spatial relationships between the objects). 2. Based on the image and scene description, provide a detailed description of the objects in the image by combining common world knowledge (including characteristics such as length, width, height, shape, color, position, and spatial relationships with other objects). Output the objects and their descriptions in this step. 3. Based on the level-one command, scene description, and object descriptions, analyze and output the final scene that needs to be achieved (including how the objects should be arranged). 4. Infer which objects need to be operated on from the level-one command by combining the image, and output the objects that need to be operated on in this step. 5. A level-three command is an instruction to perform a specific action on an object with a clear name, and it must include a precise placement location (e.g., left, right, front, back, or above another specific object), such as placing an apple on a book. Based on the level-one command and its intent, generate a clear level-three command for each object that needs to be operated on, and output the level-three commands. 6. Determine the order of execution for each level-three command by combining the image and common world knowledge to ensure that the entire process is smooth and logical. Output the sorted level-three commands. The output format should be [command1, command2, command3.....], where command1 represents the disassembled level-three commands. Strictly follow the format. Please think step by step."
extract_command3_que="Please extract the level-three commands from the above description. Do not make any changes to the level-three commands mentioned in the description. The output format should be: [command1, command2, command3....]. Strictly follow the output format (no need to use quotation marks inside the brackets, make sure to use commas instead of '，'), command1 is just a placeholder and does not need to be output. Do not output anything else."

# picture_dir = "G:\\xjtu\\llm\\project\\is_direct\\rgb2\\29.jpg"
# file_path = 'G:\\xjtu\\llm\\project\\is_direct\\is_direct.xlsx'
picture_dir = "G:\\xjtu\\llm\\project\\command\\data_new\\rgb1\\"
file_path = 'G:\\xjtu\\llm\\project\\is_direct\\is_direct.xlsx'


res_file_path = "G:\\xjtu\\llm\\project\\command\\output\\command_one_all_8_14.xlsx"

if __name__ == '__main__':
    data_init = data.copy()

    wb = Workbook()
    sheet = wb.active
    sheet.append(["image","command1","command3s","command3"])


    for i in range(2,501):
        print(str(i))
        print("---------------add query---------------")
        image_name, command1 = get_name(res_file_path,i)
        image = picture_dir + image_name


        command3_query = get_text_image(user, command2_dismantle_txt.format(command1), image)
        data_init["messages"].append(command3_query)
        res_command3 = consult_GPT(data_init)
        assistant_dismantle_res = res_command3.json()["choices"][0]["message"]["content"]
        data_init["messages"].pop()

        #提取三级指令
        data_extract = data_init
        user_extract = get_text(user, assistant_dismantle_res+extract_command3_que)
        data_extract["messages"].append(user_extract)
        res_des = consult_GPT(data_extract)
        assistant_extract_res = res_des.json()["choices"][0]["message"]["content"]
        data_extract["messages"].pop()

        command3s = convert_to_tuple(assistant_extract_res)


        # 保存结果到Excel文件
        sheet.append(
                [image.replace("G:\\xjtu\\llm\\project\\command\\data_new\\rgb1\\", ""), command1,
                 assistant_dismantle_res,assistant_extract_res])

        # 实时保存Excel文件
        wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_8_15_command1_all_dismantle.xlsx")

    # 最终保存Excel文件（可选，因为每次写入后已保存）
    wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_all_8_15_command1_all_dismantle.xlsx")








