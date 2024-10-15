import requests
import json
import base64
import pandas as pd
from openpyxl import Workbook, load_workbook
import random
import re


BUFFER_SIZE = 3
VAGUE_SIZE = 4
FILTER_SIZE = 1
url = "https://flag.smarttrot.com/v1/chat/completions"
api_secret_key = 'sk-zk2818679870614b777603fbc0bf5fc95af668bb9355b277'

headers = {
    "Content-Type": "application/json",
    'Accept': 'application/json',
    'Authorization': "Bearer " + api_secret_key
}

# "role": "assistant"/"user"
assistant = "assistant"
user = "user"

data = {
    "max_tokens": 1200,
    "model": "gpt-4o-2024-05-13",
    # "model": "gpt-4o-mini-2024-07-18",
    "temperature": 0.8,
    "top_p": 1,
    "presence_penalty": 1,
    "messages": [
        # {
        #     "role": "assistant",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "I'm going to give you 1 examples to learn from.Example1: Description:The image shows a scene with a white table on which a blue cube and an open cardboard box are placed. The blue cube is situated on the lower left side of the image, resting on the white surface of the table. It appears to be a solid object with a noticeable shadow indicating a light source from the upper right side.The cardboard box is positioned on the table, closer to the top right corner of the image. The box is open, with its flaps standing upright, and it seems to be empty. The cardboard has a typical brown color with visible creases at the edges where the flaps have been folded.In the background, there's a radiator against the wall and some cables that can be seen hanging loosely, suggesting this might be a workspace or an area where some sort of work or activity is being carried out.The overall scenario seems to be set up for an activity, possibly for the purpose of demonstrating or testing the functionality of a robot arm in picking and placing objects, considering your mention of a two-finger claw robot arm. The blue cube could be the object that needs to be manipulated, and the cardboard box could be the target location for placing the object. Object list: [blue_cube, cardboard_box].  Generated task instruction:Put the blue_cube in the cardboard_box. Is it a vague instruction:No.  Atomic action after disassembly:[(grab,blue_cube),(place,cardboard_box)]."
        #         },
        #         {
        #             "type": "image_url",
        #             "image_url": {
        #                 "url": f"data:image/jpeg;base64,{base64_image_box}"
        #             }
        #         }
        #     ]
        # },
        # {
        #     "role": "user",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "Please complete the task according to the example."
        #         }
        #     ]
        # }
    ]
}


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def read_excel_data(file_path, n, num):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    result = ""
    for row in sheet.iter_rows(min_row=num+1, max_row=num+1, values_only=True):
        print(row[n])
        result = row[n]

    wb.close()
    return result


def get_example(num, file_path):
    description_exa = "Description:" + read_excel_data(file_path, 1, num)
    print(description_exa)
    obj_list_exa = read_excel_data(file_path, 2, num)
    instruction_exa = read_excel_data(file_path, 3, num)
    is_vague_exa = read_excel_data(file_path, 4, num)
    action_list = read_excel_data(file_path, 5, num)
    return description_exa, obj_list_exa, instruction_exa, is_vague_exa, action_list


def get_text_image(role, txt, base64_image):
    base64_image = encode_image(base64_image)
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            },
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{base64_image}"
                }
            }
        ]
    }

    return res


def get_text(role, txt):
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            }
        ]
    }

    return res


def output_format(dict_data):
    for key, value in dict_data.items():
        if key == 'messages':
            print(f"{key}:")
            for item in dict_data['messages']:
                role = item['role']
                print(f"Role: {role}")
                for content_item in item['content']:
                    if isinstance(content_item, dict):
                        if content_item['type'] == 'text':
                            print(f"  Text: {content_item['text']}")
                        elif content_item['type'] == 'image_url':
                            print("  Image url: [NOT DISPLAYED]")
                        else:
                            print(f"  {content_item}")
                    else:
                        print(f"  {item['content']}")
                        break
        else:
            print(f"{key}: {value}")

# #改
# def extract_objects_from_string(s):
#     pattern = r'\[(.*?)(?:,\s*(.*?))?\]'
#     match = re.search(pattern, s)
#     if match:
#         object1 = match.group(1).strip()
#         object2 = match.group(2).strip() if match.group(2) is not None else None
#         return object1, object2
#     return None, None
def extract_objects_from_string(assistant_obj_res):
    # 找到方括号内的内容
    start = assistant_obj_res.find('[')
    end = assistant_obj_res.find(']')

    if start != -1 and end != -1:
        # 提取方括号内的内容
        objects_str = assistant_obj_res[start + 1:end]
        # 按逗号分隔并去除多余的空格
        objects = [obj.strip() for obj in objects_str.split(',')]
        return objects
    else:
        return []


def convert_to_tuple(input_str):
    # 去掉字符串的起始和结束方括号以及两端的空白字符
    trimmed_str = input_str.strip()[1:-1]
    # 按逗号分隔字符串
    split_str = trimmed_str.split(',')
    # 将分割后的列表转换成元组
    result_tuple = tuple(split_str)
    return result_tuple

def consult_GPT(question):
    response = requests.post(url, headers=headers, data=json.dumps(question))
    result = response
    # response = requests.post(url, headers=headers, data=json.dumps(question).encode('utf-8'))
    # result = response.content.decode("utf-8")
    # result = requests.post(url, json.dumps(params), headers=headers)
    # print(result.json())
    return result


def get_example_all(file_path):
    min_row = 2
    max_row = 4
    descriptions, objects, command_3s,command_4s,tasks = read_excel_data_all(file_path, min_row, max_row)
    res = ""
    for i in range(min_row, max_row + 1):

        name = "Example" + str(i - 1)+","
        task = "Task:" + tasks[i - 2] + ","
        description = "Description:" + descriptions[i-2] + ","
        object = "object:" + objects[i-2] + ","
        command_3="command:" + command_3s[i-2] + ","
        command_4 = "command_4:" + command_4s[i-2] + "."
        res = res + name + task + description + object + command_3
    return res

def read_excel_data_all(file_path, min_row, max_row):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    # names=[]
    descriptions = []
    objects=[]
    command_3s=[]
    command_4s=[]
    tasks=[]
    # result = ""
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        # print(row[n])
        # names.append(row[0])
        descriptions.append(row[1])
        objects.append(row[2])
        command_3s.append(row[3])
        command_4s.append(row[5])
        tasks.append(row[6])
        # result = row[n]

    wb.close()
    return descriptions, objects, command_3s, command_4s,tasks

# response = requests.post(url, headers=headers, data=json.dumps(data).encode('utf-8'))
# result = response.content.decode("utf-8")
#
# print(result)




# command2_que="请根据该图片，生成尽可能多的2级指令。2级指令应该包含一个概括性的对象类别（必须对应一个以上的对象）和一步明确的操作步骤。思考步骤：1、首先从图片中提取物体，并结合世界通用知识对物体进行详细的特征描述（必须包括物体的长，宽，高，颜色，形状等特征。），输出物体描述。2、从描述中提取物体列表，格式为[object1,object2,object3.....]。在这个步骤中严格按照格式输出，不需要输出别的内容。2、根据物体列表以及物体描述，结合世界通用知识，对物体进行分类，输出格式为[object1,object2,object3]（类别）,[object4,object5]（类别）。在这个步骤中需要严格按照格式输出，不需要输出别的内容。3、根据物体分类结果以及物体描述，结合世界通用知识，生成尽可能多的合适的2级指令。要注意2级指令应该包含一个概括性的对象类别（不可以写具体的物体名字）（必须对应多个对象）和一步明确的操作，可以包含逻辑性。请输出最终的二级指令，在这个步骤中只需要输出二级指令，不需要输出其他内容。以下是一些例子供参考：1、将所有的水果放到篮子里。2、把所有的文具以从大到小的顺序堆叠到书本上。3、将所有的玩具从颜色的深浅顺序收进玩具箱里。4、将所有木块放到托盘左边。请你一步一步思考。"
# extract_command2_que="请从以上描述中提取出二级指令,不需要对描述中的二级指令做任何更改，输出格式为:[command1,command2,command3....],严格按照输出格式输出(注意一定要使用,而不是，)，不需要输出其他内容"
command2_que="Please generate as many level-two commands as possible based on the image. A level-two command should include a general object category (must correspond to more than one object) and a specific action step. Thought process: 1. First, extract objects from the image and provide a detailed description of the objects based on common world knowledge (must include characteristics such as length, width, height, color, shape, etc.). Output the object descriptions. 2. Extract a list of objects from the descriptions, in the format [object1, object2, object3.....]. Strictly follow the format and do not output anything else in this step. 3. Classify the objects based on the object list and descriptions, combining common world knowledge. Output format: [object1, object2, object3] (category), [object4, object5] (category). Strictly follow the format and do not output anything else in this step. 4. Based on the object classification results and object descriptions, generate as many appropriate level-two commands as possible, combining common world knowledge. Note that a level-two command should include a general object category (specific object names should not be used) and a specific action, possibly with logic. Please output the final level-two commands. In this step, only output the level-two commands, and do not output anything else. Here are some examples for reference: 1. Put all the fruits in the basket. 2. Stack all the stationery on the book in order from large to small. 3. Arrange all the toys in the toy box in order of color from dark to light. 4. Place all the wooden blocks on the left side of the tray. Please think step by step."
extract_command2_que="Please extract the level-two commands from the above description. Do not make any changes to the level-two commands mentioned in the description. Output format: [command1, command2, command3....]. Strictly follow the output format (make sure to use commas instead of '，'), and do not output anything else."



picture_dir = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\picture\\"
vague_pic = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\vague_pic\\"
# base64_image_box = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\盒子箱子(1).jpg")
# base64_image_mouse = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\鼠标.jpg")
file_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\command2_generate_examples.xlsx'
vague_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\vague.xlsx'
filter_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\filter.xlsx'
# test_image = "E:\\科研\\大四科研\\具身智能\\周报\\03.28\\test.jpg"
# image = "E:\\data_my\\test.jpg"






if __name__ == '__main__':

    # 得到前100张图片的路径，存储在列表image_paths中
    import os
    directory_path = "G:\\xjtu\\llm\\project\\command\\data_new\\rgb1"
    image_paths = []
    for entry in os.scandir(directory_path):
        image_paths.append(entry.path)

    # 自定义排序函数，根据图片文件名进行排序
    def image_sort_key(path):
        filename = os.path.splitext(os.path.basename(path))[0]  # 提取文件名，不包括文件扩展名
        return int(filename)

    # 按指定顺序对文件路径列表进行排序
    image_paths.sort(key=image_sort_key)

    image_paths1 = image_paths[200:400]
    image_paths2 = image_paths[174:200]
    # image_paths = list(set(image_paths1 + image_paths2))
    image_paths = list(set(image_paths1))
    image_paths.sort(key=image_sort_key)
    wb = Workbook()
    sheet = wb.active
    sheet.append(["image","description","command2"])

    num = 0
    #获取全部例子
    # example_txt = get_example_all(file_path)

    for image in image_paths:
        print("!!!!!!!!!!!!!!")
        print(num)
        num=num+1
        # init
        #初始化数据
        data_init = data.copy()

        #按步骤生成物体描述、物体分类、2级指令
        # Generating descriptions
        data_command2 = data_init
        user_command2 = get_text_image(user, command2_que, image)
        data_command2["messages"].append(user_command2)
        res_des = consult_GPT(data_command2)
        assistant_commands_res = res_des.json()["choices"][0]["message"]["content"]
        data_command2["messages"].pop()


        #提取二级指令
        data_extract = data_init
        user_extract = get_text(user, assistant_commands_res+extract_command2_que)
        data_extract["messages"].append(user_extract)
        res_des = consult_GPT(data_extract)
        assistant_extract_res = res_des.json()["choices"][0]["message"]["content"]
        data_extract["messages"].pop()

        command2s = convert_to_tuple(assistant_extract_res)


        for command2 in command2s:
            print(command2)
            # 保存结果到Excel文件
            sheet.append(
                [image.replace("G:\\xjtu\\llm\\project\\command\\data_new\\rgb1\\", ""), assistant_commands_res, command2])

            # 实时保存Excel文件
            wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_8_14_command2_generate.xlsx")

    # 最终保存Excel文件（可选，因为每次写入后已保存）
    wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_all_8_14_command2_generate.xlsx")

