import requests
import json
import base64
import pandas as pd
from openpyxl import Workbook, load_workbook
import random
import re


BUFFER_SIZE = 3
VAGUE_SIZE = 4
FILTER_SIZE = 1
url = "https://flag.smarttrot.com/v1/chat/completions"
api_secret_key = ''

headers = {
    "Content-Type": "application/json",
    'Accept': 'application/json',
    'Authorization': "Bearer " + api_secret_key
}

# "role": "assistant"/"user"
assistant = "assistant"
user = "user"

data = {
    "max_tokens": 1200,
    "model": "gpt-4o-2024-05-13",
    # "model": "gpt-4o-mini-2024-07-18",
    "temperature": 0.8,
    "top_p": 1,
    "presence_penalty": 1,
    "messages": [
        # {
        #     "role": "assistant",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "I'm going to give you 1 examples to learn from.Example1: Description:The image shows a scene with a white table on which a blue cube and an open cardboard box are placed. The blue cube is situated on the lower left side of the image, resting on the white surface of the table. It appears to be a solid object with a noticeable shadow indicating a light source from the upper right side.The cardboard box is positioned on the table, closer to the top right corner of the image. The box is open, with its flaps standing upright, and it seems to be empty. The cardboard has a typical brown color with visible creases at the edges where the flaps have been folded.In the background, there's a radiator against the wall and some cables that can be seen hanging loosely, suggesting this might be a workspace or an area where some sort of work or activity is being carried out.The overall scenario seems to be set up for an activity, possibly for the purpose of demonstrating or testing the functionality of a robot arm in picking and placing objects, considering your mention of a two-finger claw robot arm. The blue cube could be the object that needs to be manipulated, and the cardboard box could be the target location for placing the object. Object list: [blue_cube, cardboard_box].  Generated task instruction:Put the blue_cube in the cardboard_box. Is it a vague instruction:No.  Atomic action after disassembly:[(grab,blue_cube),(place,cardboard_box)]."
        #         },
        #         {
        #             "type": "image_url",
        #             "image_url": {
        #                 "url": f"data:image/jpeg;base64,{base64_image_box}"
        #             }
        #         }
        #     ]
        # },
        # {
        #     "role": "user",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "Please complete the task according to the example."
        #         }
        #     ]
        # }
    ]
}


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def read_excel_data(file_path, n, num):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    result = ""
    for row in sheet.iter_rows(min_row=num+1, max_row=num+1, values_only=True):
        print(row[n])
        result = row[n]

    wb.close()
    return result


def get_example(num, file_path):
    description_exa = "Description:" + read_excel_data(file_path, 1, num)
    print(description_exa)
    obj_list_exa = read_excel_data(file_path, 2, num)
    instruction_exa = read_excel_data(file_path, 3, num)
    is_vague_exa = read_excel_data(file_path, 4, num)
    action_list = read_excel_data(file_path, 5, num)
    return description_exa, obj_list_exa, instruction_exa, is_vague_exa, action_list


def get_text_image(role, txt, base64_image):
    base64_image = encode_image(base64_image)
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            },
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{base64_image}"
                }
            }
        ]
    }

    return res


def get_text(role, txt):
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            }
        ]
    }

    return res


def output_format(dict_data):
    for key, value in dict_data.items():
        if key == 'messages':
            print(f"{key}:")
            for item in dict_data['messages']:
                role = item['role']
                print(f"Role: {role}")
                for content_item in item['content']:
                    if isinstance(content_item, dict):
                        if content_item['type'] == 'text':
                            print(f"  Text: {content_item['text']}")
                        elif content_item['type'] == 'image_url':
                            print("  Image url: [NOT DISPLAYED]")
                        else:
                            print(f"  {content_item}")
                    else:
                        print(f"  {item['content']}")
                        break
        else:
            print(f"{key}: {value}")

# #改
# def extract_objects_from_string(s):
#     pattern = r'\[(.*?)(?:,\s*(.*?))?\]'
#     match = re.search(pattern, s)
#     if match:
#         object1 = match.group(1).strip()
#         object2 = match.group(2).strip() if match.group(2) is not None else None
#         return object1, object2
#     return None, None
def extract_objects_from_string(assistant_obj_res):
    # 找到方括号内的内容
    start = assistant_obj_res.find('[')
    end = assistant_obj_res.find(']')

    if start != -1 and end != -1:
        # 提取方括号内的内容
        objects_str = assistant_obj_res[start + 1:end]
        # 按逗号分隔并去除多余的空格
        objects = [obj.strip() for obj in objects_str.split(',')]
        return objects
    else:
        return []


def convert_to_tuple(assistant_sort_res):
    # 去掉字符串两端的单引号
    assistant_sort_res = assistant_sort_res.strip("'")

    # 将字符串按分号分割成多个物体组合
    object_pairs = assistant_sort_res.split(';')

    # 将这些物体组合存储到一个元组中
    result_tuple = tuple(object_pairs)

    return result_tuple

def consult_GPT(question):
    response = requests.post(url, headers=headers, data=json.dumps(question))
    result = response
    # response = requests.post(url, headers=headers, data=json.dumps(question).encode('utf-8'))
    # result = response.content.decode("utf-8")
    # result = requests.post(url, json.dumps(params), headers=headers)
    # print(result.json())
    return result


def get_example_all(file_path):
    min_row = 2
    max_row = 4
    descriptions, objects, command_3s,command_4s,tasks = read_excel_data_all(file_path, min_row, max_row)
    res = ""
    for i in range(min_row, max_row + 1):

        name = "Example" + str(i - 1)+","
        task = "Task:" + tasks[i - 2] + ","
        description = "Description:" + descriptions[i-2] + ","
        object = "object:" + objects[i-2] + ","
        command_3="command:" + command_3s[i-2] + ","
        command_4 = "command_4:" + command_4s[i-2] + "."
        res = res + name + task + description + object + command_3
    return res

def read_excel_data_all(file_path, min_row, max_row):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    # names=[]
    descriptions = []
    objects=[]
    command_3s=[]
    command_4s=[]
    tasks=[]
    # result = ""
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        # print(row[n])
        # names.append(row[0])
        descriptions.append(row[1])
        objects.append(row[2])
        command_3s.append(row[3])
        command_4s.append(row[5])
        tasks.append(row[6])
        # result = row[n]

    wb.close()
    return descriptions, objects, command_3s, command_4s,tasks

# response = requests.post(url, headers=headers, data=json.dumps(data).encode('utf-8'))
# result = response.content.decode("utf-8")
#
# print(result)

goal_txt = "Goal:My ultimate goal is for you to generate the task instructions suitable for the two-finger claw robot arm according to the pictures I provided, and disassemble them into the format I required. This tip is only for you to learn, and there is no need to give the result."
robot_constraint = "Robot restrictions: the target of the robot can not be fragile objects such as glass and display, but objects such as apples and bananas can be grasped and are not fragile objects. At the same time, it is not allowed to grasp human parts and perform tasks that may harm humans."
des_txt = "I'm going to give you 1 examples to learn from.Example1: Description:The image shows a scene with a white table on which a blue cube and an open cardboard box are placed. The blue cube is situated on the lower left side of the image, resting on the white surface of the table. It appears to be a solid object with a noticeable shadow indicating a light source from the upper right side.The cardboard box is positioned on the table, closer to the top right corner of the image. The box is open, with its flaps standing upright, and it seems to be empty. The cardboard has a typical brown color with visible creases at the edges where the flaps have been folded.In the background, there's a radiator against the wall and some cables that can be seen hanging loosely, suggesting this might be a workspace or an area where some sort of work or activity is being carried out.The overall scenario seems to be set up for an activity, possibly for the purpose of demonstrating or testing the functionality of a robot arm in picking and placing objects, considering your mention of a two-finger claw robot arm. The blue cube could be the object that needs to be manipulated, and the cardboard box could be the target location for placing the object."
#生成物体信息
des_que = "Please describe the objects on the desk in detail based on the image. The description must include the length, width, height, shape, and thickness of the objects, as well as whether the objects are lying horizontally or standing vertically, using both the image and general world knowledge."
#"请根据图像详细描述桌子上的物体。描述必须包括物体的长度、宽度、高度、形状和厚度，以及物体是水平放置还是垂直放置，结合图像和常识进行描述。"
#change
obj_que = "Based on the object descriptions and images, extract all objects suitable for grasping and placing by a two-finger gripper arm. The objects to be grasped must be suitable for grasping by a UR3 two-finger gripper arm (suitable grasping features: the thickness on the table must be greater than 3 cm, and the maximum width in the grasping direction must be less than 8 cm). The objects to be placed must be suitable for placing, and the objects to be grasped should not already be above the objects to be placed. The reference example is only for reference, the extracted objects must be mentioned in the object description). Output format (do not omit punctuation): [object1, object2, object3...],[object5]. The first list of objects is suitable for grasping, and the second list of objects is suitable for placing. The output should only contain data in this format, without any additional prompts. Example of thickness comparison: 2 cm is greater than 0.3 cm, 2.5 cm is greater than 2.15 cm. The thickness of objects suitable for grasping must be greater than three centimeters."
#obj_que"根据物体描述和图像，提取所有适合两指抓手操作的抓取和放置的物体。要抓取的物体必须适合UR3两指抓手臂抓取（适合抓取的特征：桌上的厚度必须大于3厘米，抓取方向的最大宽度必须小于8厘米）。要放置的物体必须适合放置，要抓取的物体不应已经位于要放置的物体之上。参考示例仅供参考，提取的物体必须在物体描述中提到。输出格式（不要省略标点）：[object1, object2, object3...],[object5]。第一个列表的物体适合抓取，第二个列表的物体适合放置。输出应仅包含这种格式的数据，不要有任何额外的提示。厚度比较示例：2 cm 大于 0.3 cm，2.5 cm 大于 2.15 cm。适合抓取的物体厚度必须大于三厘米。"
ins_que = "Based on the examples of different tasks mentioned above (including pick-and-place tasks, stacking tasks, and tasks involving placing objects to the left, right, front, or back of another object), and integrating general world knowledge, please select an appropriate task for two objects in the obj_list from the image and generate a task instruction suitable for a two-finger gripper. The objects to be manipulated must be in the obj_list, and their names must match exactly. The format of obj_list is '[object1, object2]'. Object1 is the object to be picked up, and object1 must be placed in front of, behind, to the left of, to the right of, or on top of object2. Please specify this in the command. Ensure that the objects to be manipulated are in the obj_list, and their names must match exactly. The task instruction does not need to list the objects (obj_list). Please think step by step and output the task instruction, but only the task instruction is needed, without explanation or additional content."
#根据上面对话中不同任务的示例(分别有放置抓取任务、堆叠任务、放在物体左右前后方位的任务)，并结合通用世界知识，请为图像中的 obj_list 中的两个物体选择合适的任务并生成适合两指抓手操作的任务指令。要操作的物体必须在 obj_list 中，并且名称必须完全相同。obj_list 的格式为 '[object1,object2]'。Object1 是要抓取的物体，object1必须是放置在object2的前后左右或上边，请你在命令中具体指出。请确保要操作的物体必须在 obj_list 中，并且名称必须完全相同。任务指令不需要分点。任务指令不需要输出对象列表 (obj_list)。请一步一步思考，输出任务指令，最后只需要输出任务指令，并不需要解释和其他内容。
vague_que = "Is this instruction a fuzzy instruction? The fuzzy instruction means that this instruction is very abstract and can be broken down into a series of actions. It is a high-level instruction, not all composed of atomic actions of the manipulator, not similar to grasp and place, which is suitable for the operation of the two-finger gripper manipulator arm, or the object of operation is not suitable for the operation of the two-finger gripper manipulator arm. If it is composed of multiple atomic actions suitable for the operation of the two-finger gripper manipulator, it does not belong to the fuzzy instruction. The answer can be yes or no and nothing else. The instructions for this judgment are:  "
disassemble_que = "Please generate the action sequence that can complete the task according to the task instructions and scenes. The actions only include grabbing and placing. The action sequence template is as follows: <(action1, obj1), (action2, obj2, Direction)>. Direction can be front, back, left, right, or above. Note that only the result of the disassembly is needed, without any other prompts or other content. The objects to be operated must be the ones extracted above, and their names must be the same."
#请根据任务指令和场景生成可以完成任务的动作序列。动作只包括抓取和放置。动作序列模板如下：<(action1,obj1),(action2,obj2,Direction)>。Direction可以是front、back、left、right、above。注意只需要拆解后的结果，不需要其他提示或其他内容。要操作的物体必须是上面提取的物体，名称必须相同。
dis_init_que = "This task instruction is a fuzzy instruction, which is not only composed of atomic actions such as place, grab, etc. This instruction can be further decomposed into a series of task instructions composed of atomic actions. The generated results only include the decomposed results, and no other prompt content is allowed. Such as <put the apple on the plate>< move the rag >. The vague instructions that need further unpacking are as follows:"
sort_que = "Based on the description, image, and the list of objects, please generate pairs of objects suitable for the operation of the two-finger gripper arm in the image. Each pair should consist of two objects from the list of objects, with names exactly matching, formatted as '[object1,object2]'. Object1 is the object to be grasped, and object2 is the location where object1 is to be placed. As a reference example: object2 can be a tray (corresponding to placing object1 in the tray), a book (corresponding to placing object1 on the book), an apple (corresponding to placing object1 to the right of the apple), or a wooden block (corresponding to stacking object1 on the wooden block). There may be multiple such pairs of objects in one image. Finally, evaluate the reasonableness of executing the object pairs and store all pairs of objects together in order from most to least reasonable, in the format: [object1,object2];[object3,object4]; etc. Please note that punctuation marks should not be omitted.Only output in the required format, no additional explanations needed."
#sort_que"根据描述、图像和物体列表，请生成适合图像中两指抓手操作的物体对。每对应由物体列表中的两个物体组成，名称必须完全匹配，格式为 '[object1,object2]'。Object1 是要抓取的物体，object2 是 object1 要被放置的位置。作为参考示例：object2 可以是托盘（对应将 object1 放入托盘中）、书（对应将 object1 放在书上）、苹果（对应将 object1 放置在苹果右边）或木块（对应将 object1 堆叠在木块上）。一张图像中可能有多个这样的物体对。最后，评估执行这些物体对的合理性，并按从最合理到最不合理的顺序存储所有物体对，格式为：[object1,object2];[object3,object4];等。请注意，不要省略标点符号。只按要求格式输出，不需要额外解释。"
exa_prefix = "I'm going to give you 1 examples to learn from.Example1: "
des_prefix = " The descriptions of the objects in the image are: "
obj_prefix = " The operable objects are: "
ins_prefix = " The obj_list is: "

sort_prefix = "The list of actionable objects is:"

picture_dir = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\picture\\"
vague_pic = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\vague_pic\\"
# base64_image_box = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\盒子箱子(1).jpg")
# base64_image_mouse = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\鼠标.jpg")
file_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\command3_generate_examples.xlsx'
vague_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\vague.xlsx'
filter_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\filter.xlsx'
# test_image = "E:\\科研\\大四科研\\具身智能\\周报\\03.28\\test.jpg"
# image = "E:\\data_my\\test.jpg"






if __name__ == '__main__':

    # 得到前100张图片的路径，存储在列表image_paths中
    import os
    directory_path = "G:\\xjtu\\llm\\project\\command\\data_new\\rgb1"
    image_paths = []
    for entry in os.scandir(directory_path):
        image_paths.append(entry.path)

    # 自定义排序函数，根据图片文件名进行排序
    def image_sort_key(path):
        filename = os.path.splitext(os.path.basename(path))[0]  # 提取文件名，不包括文件扩展名
        return int(filename)

    # 按指定顺序对文件路径列表进行排序
    image_paths.sort(key=image_sort_key)

    image_paths1 = image_paths[0:200]
    image_paths2 = image_paths[174:200]

    # image_paths = list(set(image_paths1 + image_paths2))
    image_paths = list(set(image_paths1))
    image_paths.sort(key=image_sort_key)
    wb = Workbook()
    sheet = wb.active
    sheet.append(["image","assistant_obj_res","assistant_ins_res", "assistant_is_vague_res", "assistant_disassemble_res"])

    num = 0
    #获取全部例子
    example_txt = get_example_all(file_path)

    for image in image_paths:
        print("!!!!!!!!!!!!!!")
        print(num)
        num=num+1
        # init
        #初始化数据
        data_init = data.copy()
        goal = get_text(user, goal_txt)
        robot_cons = get_text(user, robot_constraint)
        data_init["messages"].append(goal)
        data_init["messages"].append(robot_cons)
        output_format(data_init)


        #生成物体描述
        # Generating descriptions
        data_des = data_init
        user_des = get_text_image(user, des_que, image)
        data_des["messages"].append(user_des)
        output_format(data_des)
        res_des = consult_GPT(data_des)
        assistant_des_res = res_des.json()["choices"][0]["message"]["content"]
        output_format(data_des)
        data_des["messages"].pop()



        # 提取可操作的对象
        data_obj = data_des
        obj_user = get_text(user, obj_que + des_prefix + assistant_des_res)
        data_obj["messages"].append(obj_user)
        res_obj = consult_GPT(data_obj)
        assistant_obj_res = res_obj.json()["choices"][0]["message"]["content"]
        #问题：同一图片每次提取不一致
        # 找到第一个和第二个方括号的内容
        first_list_content = assistant_obj_res[assistant_obj_res.find('[') + 1: assistant_obj_res.find(']')]
        second_list_content = assistant_obj_res[assistant_obj_res.rfind('[') + 1: assistant_obj_res.rfind(']')]

        # 将提取的字符串转换为列表
        grab_list = [item.strip() for item in first_list_content.split(',')]
        place_list = [item.strip() for item in second_list_content.split(',')]

        objects =  list(set(grab_list + place_list))
        print("Objects:", objects)
        data_obj["messages"].pop()


        #选出两两组合，按照易执行程度排序
        data_sort =  data_obj
        sort_user = get_text(user, sort_que + sort_prefix + assistant_obj_res)
        data_sort["messages"].append(sort_user)
        res_sort = consult_GPT(data_sort)
        assistant_sort_res = res_sort.json()["choices"][0]["message"]["content"]
        sort_result = convert_to_tuple(assistant_sort_res)
        data_sort["messages"].pop()


        for object_item in sort_result:
            num = random.randint(1, BUFFER_SIZE)
            #获取例子
            description_exa, obj_list_exa, instruction_exa, is_vague_exa, action_list = get_example(num, file_path)
            base64_image = picture_dir + str(num) + ".jpg"

            # Generating instruction
            # 生成任务指令
            data_instruction = data_obj
            # add examples (choose)

            # example_instruction_user = get_text_image(user, ins_que, base64_image)
            # example_instruction_assistant = get_text(assistant, instruction_exa)
            # data_instruction["messages"].append(example_instruction_user)
            # data_instruction["messages"].append(example_instruction_assistant)
            #放入例子
            example = get_text(user, example_txt)
            data_instruction["messages"].append(example)
            # add que, des and obj
            # assistant_obj_res可操作物体，str'[roll of tape, small black circular object]'
            ins_user = get_text_image(user, ins_que + des_prefix + assistant_des_res + obj_prefix + object_item,
                                      image)
            data_instruction["messages"].append(ins_user)
            res_ins = consult_GPT(data_instruction)
            assistant_ins_res = res_ins.json()["choices"][0]["message"]["content"]
            output_format(data_instruction)
            data_instruction["messages"].pop()
            data_instruction["messages"].pop()
            # data_instruction["messages"].pop()
            # data_instruction["messages"].pop()


            # is vague?
            # 判断任务指令是否模糊
            data_is_vague = data_instruction
            # 读取Excel表格
            df = pd.read_excel(vague_path, sheet_name='vague')
            # add examples (choose)
            for index, row in df.head(VAGUE_SIZE).iterrows():
                ins = row.iloc[0]  # 第一列保存为ins变量
                is_vague = row.iloc[1]  # 第二列保存为is_vague变量
                example_vague_user = get_text(user, vague_que + ins)
                example_vague_assistant = get_text(assistant, is_vague)
                data_is_vague["messages"].append(example_vague_user)
                data_is_vague["messages"].append(example_vague_assistant)
            # add que
            vague_user = get_text(user, vague_que + assistant_ins_res)
            data_is_vague["messages"].append(vague_user)
            res_is_vague = consult_GPT(data_is_vague)
            assistant_is_vague_res = res_is_vague.json()["choices"][0]["message"]["content"]
            output_format(data_is_vague)
            print(assistant_ins_res)
            print(assistant_is_vague_res)
            data_is_vague["messages"].pop()
            for i in range(1, VAGUE_SIZE + 1):
                data_is_vague["messages"].pop()
                data_is_vague["messages"].pop()

            # distinct instructions
            # 分解任务指令
            data_disassemble = data_is_vague
            # add examples (choose)
            example_disassemble_user = get_text_image(user,
                                                      disassemble_que + des_prefix + description_exa + obj_prefix + obj_list_exa + ins_prefix + instruction_exa,
                                                      base64_image)
            example_disassemble_assistant = get_text(assistant, action_list)
            data_disassemble["messages"].append(example_disassemble_user)
            data_disassemble["messages"].append(example_disassemble_assistant)
            # add que
            disassemble_user = get_text_image(user,
                                              disassemble_que + des_prefix + assistant_des_res + obj_prefix + object_item + ins_prefix + assistant_ins_res,
                                              image)
            data_disassemble["messages"].append(disassemble_user)
            res_disassemble = consult_GPT(data_disassemble)
            assistant_disassemble_res = res_disassemble.json()["choices"][0]["message"]["content"]
            output_format(data_disassemble)
            print(assistant_disassemble_res)
            data_disassemble["messages"].pop()
            data_disassemble["messages"].pop()
            data_disassemble["messages"].pop()

            # 保存结果到Excel文件
            sheet.append(
                [image.replace("G:\\xjtu\\llm\\project\\command\\data_new\\rgb1\\", ""), object_item, assistant_ins_res,
                 assistant_is_vague_res, assistant_disassemble_res])

            # 实时保存Excel文件
            wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_8_14_command3_generate.xlsx")

    # 最终保存Excel文件（可选，因为每次写入后已保存）
    wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_all_8_14_command3_generate.xlsx")

