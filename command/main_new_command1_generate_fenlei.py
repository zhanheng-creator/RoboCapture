# -*-coding:utf-8-*-

import requests
import json
import base64
import pandas as pd
from openpyxl import Workbook, load_workbook
import random
import re


BUFFER_SIZE = 3
VAGUE_SIZE = 4
FILTER_SIZE = 1
url = "https://flag.smarttrot.com/v1/chat/completions"
api_secret_key = ''

headers = {
    "Content-Type": "application/json",
    'Accept': 'application/json',
    'Authorization': "Bearer " + api_secret_key
}

# "role": "assistant"/"user"
assistant = "assistant"
user = "user"

data = {
    "max_tokens": 1200,
    "model": "gpt-4o-2024-05-13",
    # "model": "gpt-4o-mini-2024-07-18",
    "temperature": 0.8,
    "top_p": 1,
    "presence_penalty": 1,
    "messages": [
        # {
        #     "role": "assistant",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "I'm going to give you 1 examples to learn from.Example1: Description:The image shows a scene with a white table on which a blue cube and an open cardboard box are placed. The blue cube is situated on the lower left side of the image, resting on the white surface of the table. It appears to be a solid object with a noticeable shadow indicating a light source from the upper right side.The cardboard box is positioned on the table, closer to the top right corner of the image. The box is open, with its flaps standing upright, and it seems to be empty. The cardboard has a typical brown color with visible creases at the edges where the flaps have been folded.In the background, there's a radiator against the wall and some cables that can be seen hanging loosely, suggesting this might be a workspace or an area where some sort of work or activity is being carried out.The overall scenario seems to be set up for an activity, possibly for the purpose of demonstrating or testing the functionality of a robot arm in picking and placing objects, considering your mention of a two-finger claw robot arm. The blue cube could be the object that needs to be manipulated, and the cardboard box could be the target location for placing the object. Object list: [blue_cube, cardboard_box].  Generated task instruction:Put the blue_cube in the cardboard_box. Is it a vague instruction:No.  Atomic action after disassembly:[(grab,blue_cube),(place,cardboard_box)]."
        #         },
        #         {
        #             "type": "image_url",
        #             "image_url": {
        #                 "url": f"data:image/jpeg;base64,{base64_image_box}"
        #             }
        #         }
        #     ]
        # },
        # {
        #     "role": "user",
        #     "content": [
        #         {
        #             "type": "text",
        #             "text": "Please complete the task according to the example."
        #         }
        #     ]
        # }
    ]
}


def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def read_excel_data(file_path, n, num):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    result = ""
    for row in sheet.iter_rows(min_row=num+1, max_row=num+1, values_only=True):
        print(row[n])
        result = row[n]

    wb.close()
    return result


def get_example(num, file_path):
    description_exa = "Description:" + read_excel_data(file_path, 1, num)
    print(description_exa)
    obj_list_exa = read_excel_data(file_path, 2, num)
    instruction_exa = read_excel_data(file_path, 3, num)
    is_vague_exa = read_excel_data(file_path, 4, num)
    action_list = read_excel_data(file_path, 5, num)
    return description_exa, obj_list_exa, instruction_exa, is_vague_exa, action_list


def get_text_image(role, txt, base64_image):
    base64_image = encode_image(base64_image)
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            },
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{base64_image}"
                }
            }
        ]
    }

    return res


def get_text(role, txt):
    res = {
        "role": role,
        "content": [
            {
                "type": "text",
                "text": txt
            }
        ]
    }

    return res


def output_format(dict_data):
    for key, value in dict_data.items():
        if key == 'messages':
            print(f"{key}:")
            for item in dict_data['messages']:
                role = item['role']
                print(f"Role: {role}")
                for content_item in item['content']:
                    if isinstance(content_item, dict):
                        if content_item['type'] == 'text':
                            print(f"  Text: {content_item['text']}")
                        elif content_item['type'] == 'image_url':
                            print("  Image url: [NOT DISPLAYED]")
                        else:
                            print(f"  {content_item}")
                    else:
                        print(f"  {item['content']}")
                        break
        else:
            print(f"{key}: {value}")

# #改
# def extract_objects_from_string(s):
#     pattern = r'\[(.*?)(?:,\s*(.*?))?\]'
#     match = re.search(pattern, s)
#     if match:
#         object1 = match.group(1).strip()
#         object2 = match.group(2).strip() if match.group(2) is not None else None
#         return object1, object2
#     return None, None
def extract_objects_from_string(assistant_obj_res):
    # 找到方括号内的内容
    start = assistant_obj_res.find('[')
    end = assistant_obj_res.find(']')

    if start != -1 and end != -1:
        # 提取方括号内的内容
        objects_str = assistant_obj_res[start + 1:end]
        # 按逗号分隔并去除多余的空格
        objects = [obj.strip() for obj in objects_str.split(',')]
        return objects
    else:
        return []


def convert_to_tuple(input_str):
    # 去掉字符串的起始和结束方括号以及两端的空白字符
    trimmed_str = input_str.strip()[1:-1]
    # 按逗号分隔字符串
    split_str = trimmed_str.split(',')
    # 将分割后的列表转换成元组
    result_tuple = tuple(split_str)
    return result_tuple

def consult_GPT(question):
    response = requests.post(url, headers=headers, data=json.dumps(question))
    result = response
    # response = requests.post(url, headers=headers, data=json.dumps(question).encode('utf-8'))
    # result = response.content.decode("utf-8")
    # result = requests.post(url, json.dumps(params), headers=headers)
    # print(result.json())
    return result


def get_example_all(file_path):
    min_row = 4
    max_row = 8
    descriptions, command_1s, tasks = read_excel_data_all(file_path, min_row, max_row)
    res = ""
    for i in range(min_row, max_row + 1):

        name = "例子" + str(i - min_row+1)+","
        description = "场景描述:" + descriptions[i - min_row] + ","
        command_1="一级指令:" + command_1s[i-min_row] + ","
        task = "任务类型:" + tasks[i - min_row] + ","

        res = res + name + description + command_1 + task
    return res

def read_excel_data_all(file_path, min_row, max_row):
    # 打开Excel文件
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active

    # names=[]
    descriptions = []
    command_1s=[]
    tasks=[]

    # result = ""
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        # print(row[n])
        # names.append(row[0])
        descriptions.append(row[1])
        command_1s.append(row[2])
        tasks.append(row[3])


    wb.close()
    return descriptions, command_1s, tasks

# response = requests.post(url, headers=headers, data=json.dumps(data).encode('utf-8'))
# result = response.content.decode("utf-8")
#
# print(result)

#7.30日的提示词
#三级指令是对具有明确物体名称的物体进行明确操作的指令，同时三级指令应具有一个明确的放置位置（可以放置到别的物体的左右前后上边），如将苹果放置到书本上。针对该图片的一级指令是{对桌面上的物体进行分类}，请你将其分解为三级指令。步骤：1、结合图片分析当前场景，并输出当前场景的详细描述（包括有什么物体、物体的位置、物体之间的空间关系）。2、根据图片和场景描述，结合世界通用知识对图中的物体的特征进行详细描述（包括长、宽、高、形状、颜色、位置、与其他物体的空间关系等特征），在该步骤中输出物体及其描述。3、结合图片从一级指令中推理出现在需要对哪些物体进行操作,在该步骤中输出需要操作的物体。4、根据分解后的操作步骤，为每个具体物体生成一个明确的三级指令，输出三级指令。5、结合图片和通用世界知识，确定每个三级指令执行的顺序，使得整个过程流畅且合理。输出经过排序后的三级指令。输出格式为[command1,command2,command3.....],command1代指分解出的三级指令，严格按照格式输出。请你一步一步思考。

# command1_que="一级指令是对象不明确（不应该包含需要抓取的物体类别以及需要抓取的具体物体名字）或者操作不明确的指令。例子有：对桌面上的物体进行分类、将桌面上与盒子里的物体同类别的其他物体放进盒子里、将桌面上与苹果同类的物体都放到盒子前方、将桌面上可堆叠的物体堆叠在一起。生成一级指令的思考步骤：1、提取物体及其特征：从图片中提取出所有物体，并结合图片和世界通用知识描述物体的属性特征（必须包括物体的情况、颜色、位置），在该步骤中输出提取出的物体名字以及物体详细的特征描述。2、对提取出的物体进行分类。3、根据图片和物体描述，结合通用知识生成一条一级指令。请你一步一步思考。"
# extract_command1_que="请从以上描述中提取出一级指令,不需要对描述中的一级指令做任何更改，只需要输出一级指令，不需要输出其他内容"
command1_que="A level-one command is an instruction where the object is not clearly specified (it should not include the category of objects to be grabbed or the specific name of the objects to be grabbed) or the operation is unclear. Examples include: classifying the objects on the table, placing other objects of the same category as the ones in the box into the box, placing all objects of the same category as the apple in front of the box, stacking stackable objects on the table together. The thought process for generating a level-one command: 1. Extract objects and their features: Extract all objects from the image and describe the attributes and characteristics of the objects by combining the image and common knowledge (must include the condition, color, and location of the objects). In this step, output the names of the extracted objects and a detailed description of the objects' characteristics. 2. Classify the extracted objects. 3. Based on the image and object descriptions, generate a level-one command by combining common knowledge. Please think step by step."
extract_command1_que="Please extract the level-one command from the above description. You don't need to make any changes to the level-one command mentioned in the description; just output the level-one command without any additional content."


picture_dir = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\picture\\"
vague_pic = "G:\\xjtu\\llm\\project\\command\\data_my\\data\\vague_pic\\"
# base64_image_box = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\盒子箱子(1).jpg")
# base64_image_mouse = encode_image("E:\\科研\\大四科研\\具身智能\\周报\\03.28\\鼠标.jpg")
file_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\一级指令.xlsx'
vague_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\vague.xlsx'
filter_path = 'G:\\xjtu\\llm\\project\\command\\data_my\\filter.xlsx'
# test_image = "E:\\科研\\大四科研\\具身智能\\周报\\03.28\\test.jpg"
# image = "E:\\data_my\\test.jpg"






if __name__ == '__main__':

    # 得到前100张图片的路径，存储在列表image_paths中
    import os
    directory_path = "G:\\xjtu\\llm\\project\\command\\data_new\\rgb1"
    image_paths = []
    for entry in os.scandir(directory_path):
        image_paths.append(entry.path)

    # 自定义排序函数，根据图片文件名进行排序
    def image_sort_key(path):
        filename = os.path.splitext(os.path.basename(path))[0]  # 提取文件名，不包括文件扩展名
        return int(filename)

    # 按指定顺序对文件路径列表进行排序
    image_paths.sort(key=image_sort_key)

    image_paths1 = image_paths[0:200]
    image_paths2 = image_paths[174:200]
    # image_paths = list(set(image_paths1 + image_paths2))
    image_paths = list(set(image_paths1))
    image_paths.sort(key=image_sort_key)
    wb = Workbook()
    sheet = wb.active
    sheet.append(["image","description","command1"])

    num = 0
    #获取全部例子
    example_txt = get_example_all(file_path)

    for image in image_paths:
        print("!!!!!!!!!!!!!!")
        print(num)
        num=num+1
        # init
        #初始化数据
        data_init = data.copy()



        #按步骤生成物体描述、、1级指令
        # Generating descriptions
        data_command1 = data_init

        # 放入例子
        # example = get_text(user, example_txt)
        # data_command1["messages"].append(example)

        user_command1 = get_text_image(user, command1_que, image)
        data_command1["messages"].append(user_command1)
        res_des = consult_GPT(data_command1)
        assistant_commands_res = res_des.json()["choices"][0]["message"]["content"]
        data_command1["messages"].pop()
        # data_command1["messages"].pop()


        #提取1级指令
        data_extract = data_init
        user_extract = get_text(user, assistant_commands_res+extract_command1_que)
        data_extract["messages"].append(user_extract)
        res_des = consult_GPT(data_extract)
        assistant_extract_res = res_des.json()["choices"][0]["message"]["content"]
        data_extract["messages"].pop()

        # command1s = convert_to_tuple(assistant_extract_res)


        # for command1 in command1s:
        print(assistant_extract_res)
        # 保存结果到Excel文件
        sheet.append(
            [image.replace("G:\\xjtu\\llm\\project\\command\\data_new\\rgb1\\", ""), assistant_commands_res, assistant_extract_res])

        # 实时保存Excel文件
        wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_8_14_command1_generate_fenlei.xlsx")

    # 最终保存Excel文件（可选，因为每次写入后已保存）
    wb.save("G:\\xjtu\\llm\\project\\command\\output\\output_all_8_14_command1_generate_fenlei.xlsx")

